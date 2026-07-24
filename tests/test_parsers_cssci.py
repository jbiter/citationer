"""Edge-case coverage tests for the CSSCI parser."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

import openpyxl

from citationer.models.record import DocType, Institution
from citationer.parsers.cssci import CssciParser


class TestCssciParserSource:
    def test_source_name(self):
        assert CssciParser().source_name == "CSSCI"


class TestCssciDetect:
    def test_detect_rejects_unsupported_suffix(self, tmp_path: Path):
        p = tmp_path / "cssci.pdf"
        p.write_text("not cssci", encoding="utf-8")
        assert CssciParser().detect(p) is False

    def test_detect_txt_with_markers(self, tmp_path: Path):
        p = tmp_path / "cssci.txt"
        p.write_text("来源篇名\t来源作者\t期刊名称\t年份\n", encoding="utf-8")
        assert CssciParser().detect(p) is True

    def test_detect_csv_with_markers(self, tmp_path: Path):
        p = tmp_path / "cssci.csv"
        p.write_text("来源篇名,来源作者,期刊名称,年份\n", encoding="utf-8")
        assert CssciParser().detect(p) is True

    def test_detect_xlsx_with_markers(self, tmp_path: Path):
        p = tmp_path / "cssci.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["来源篇名", "来源作者", "期刊名称", "年份"])
        wb.save(p)
        assert CssciParser().detect(p) is True

    def test_detect_xlsx_empty_workbook_returns_false(self, tmp_path: Path):
        p = tmp_path / "empty.xlsx"
        openpyxl.Workbook().save(p)
        assert CssciParser().detect(p) is False

    def test_detect_xlsx_only_one_marker_returns_false(self, tmp_path: Path):
        p = tmp_path / "cssci.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["来源篇名", "其他"])
        wb.save(p)
        assert CssciParser().detect(p) is False

    def test_detect_no_active_sheet_returns_false(self, tmp_path: Path, monkeypatch):
        fake_wb = MagicMock()
        fake_wb.active = None
        monkeypatch.setattr(
            "openpyxl.load_workbook", lambda *args, **kwargs: fake_wb
        )
        p = tmp_path / "fake.xlsx"
        p.write_text("", encoding="utf-8")
        assert CssciParser().detect(p) is False


class TestCssciParseText:
    def test_parse_tab_delimited(self, tmp_path: Path):
        p = tmp_path / "cssci.txt"
        p.write_text(
            "来源篇名\t来源作者\t期刊名称\t年份\t机构\t基金\t关键词\t被引频次\n"
            "标题甲\t作者甲；作者乙\tCSSCI期刊\t2024\t清华；北大\t基金A；基金B\t关键词1；关键词2\t10\n",
            encoding="utf-8",
        )
        records = CssciParser().parse(p)
        assert len(records) == 1
        r = records[0]
        assert r.title == "标题甲"
        assert r.year == 2024
        assert len(r.authors) == 2
        assert r.authors[0].full_name == "作者甲"
        assert len(r.institutions) == 2
        assert r.institutions[0].name == "清华"
        assert r.funding == ["基金A", "基金B"]
        assert "关键词1" in r.keywords
        assert r.citation_count == 10

    def test_parse_comma_delimited(self, tmp_path: Path):
        p = tmp_path / "cssci.csv"
        p.write_text(
            "来源篇名,来源作者,期刊名称,年份\n"
            "标题乙,作者丙,CSSCI期刊,2023\n",
            encoding="utf-8",
        )
        records = CssciParser().parse(p)
        assert len(records) == 1
        assert records[0].title == "标题乙"
        assert records[0].year == 2023

    def test_parse_text_empty_returns_empty(self, tmp_path: Path):
        p = tmp_path / "empty.txt"
        p.write_text("", encoding="utf-8")
        assert CssciParser().parse(p) == []

    def test_parse_text_blank_rows_skipped(self, tmp_path: Path):
        p = tmp_path / "cssci.txt"
        p.write_text(
            "来源篇名\t来源作者\t期刊名称\t年份\n"
            "第一\t甲\t期刊\t2024\n"
            "\t\t\t\n"
            "第二\t乙\t期刊\t2023\n",
            encoding="utf-8",
        )
        records = CssciParser().parse(p)
        assert len(records) == 2

    def test_parse_alternate_headers(self, tmp_path: Path):
        p = tmp_path / "cssci.txt"
        p.write_text(
            "来源篇名\t作者\t作者机构\t期刊名称\t英文刊名\t年\t卷\t期\t页码\n"
            "中文标题\t张三\t清华大学\t期刊\tJournal\t2022\t10\t2\t1-10\n",
            encoding="utf-8",
        )
        records = CssciParser().parse(p)
        assert len(records) == 1
        r = records[0]
        assert r.title == "中文标题"
        assert r.authors[0].full_name == "张三"
        assert r.institutions[0].name == "清华大学"
        assert r.journal == "期刊"
        assert r.journal_en == "Journal"
        assert r.year == 2022
        assert r.volume == "10"
        assert r.issue == "2"
        assert r.pages == "1-10"


class TestCssciParseXlsx:
    def _save_xlsx(self, path: Path, rows: list[list[str]]) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        wb.save(path)

    def test_parse_xlsx_full_record(self, tmp_path: Path):
        p = tmp_path / "cssci.xlsx"
        self._save_xlsx(
            p,
            [
                ["来源篇名", "来源作者", "期刊名称", "年份", "关键词", "摘要"],
                ["中文标题", "作者甲；作者乙", "CSSCI期刊", "2024", "关键词1；关键词2", "摘要内容"],
            ],
        )
        records = CssciParser().parse(p)
        assert len(records) == 1
        r = records[0]
        assert r.title == "中文标题"
        assert r.year == 2024
        assert r.language == "zh"
        assert r.abstract == "摘要内容"
        assert r.doc_type == DocType.ARTICLE

    def test_parse_xlsx_empty_returns_empty(self, tmp_path: Path):
        p = tmp_path / "empty.xlsx"
        openpyxl.Workbook().save(p)
        assert CssciParser().parse(p) == []

    def test_parse_xlsx_blank_row_skipped(self, tmp_path: Path):
        p = tmp_path / "cssci.xlsx"
        self._save_xlsx(
            p,
            [
                ["来源篇名", "来源作者", "期刊名称", "年份"],
                ["第一", "甲", "期刊", "2024"],
                [None, None, None, None],
                ["第二", "乙", "期刊", "2023"],
            ],
        )
        records = CssciParser().parse(p)
        assert len(records) == 2

    def test_parse_xlsx_no_active_sheet(self, tmp_path: Path, monkeypatch):
        fake_wb = MagicMock()
        fake_wb.active = None
        monkeypatch.setattr(
            "openpyxl.load_workbook", lambda *args, **kwargs: fake_wb
        )
        p = tmp_path / "fake.xlsx"
        p.write_text("", encoding="utf-8")
        assert CssciParser().parse(p) == []


class TestCssciRowParsing:
    def test_invalid_year(self, tmp_path: Path):
        p = tmp_path / "cssci.txt"
        p.write_text(
            "来源篇名\t来源作者\t期刊名称\t年份\n"
            "标题\t甲\t期刊\t无年份\n",
            encoding="utf-8",
        )
        assert CssciParser().parse(p)[0].year is None

    def test_invalid_citation_count(self, tmp_path: Path):
        p = tmp_path / "cssci.txt"
        p.write_text(
            "来源篇名\t来源作者\t期刊名称\t年份\t被引频次\n"
            "标题\t甲\t期刊\t2024\tN/A\n",
            encoding="utf-8",
        )
        assert CssciParser().parse(p)[0].citation_count is None

    def test_authors_chinese_separators(self, tmp_path: Path):
        p = tmp_path / "cssci.txt"
        p.write_text(
            "来源篇名\t来源作者\t期刊名称\t年份\n"
            "标题\t作者甲；作者乙，作者丙\t期刊\t2024\n",
            encoding="utf-8",
        )
        r = CssciParser().parse(p)[0]
        assert [a.full_name for a in r.authors] == ["作者甲", "作者乙", "作者丙"]

    def test_keywords_filter_short_tokens(self, tmp_path: Path):
        p = tmp_path / "cssci.txt"
        p.write_text(
            "来源篇名\t来源作者\t期刊名称\t年份\t关键词\n"
            "标题\t甲\t期刊\t2024\tA；关键词1；B\n",
            encoding="utf-8",
        )
        assert CssciParser().parse(p)[0].keywords == ["关键词1"]

    def test_institutions_and_funding(self, tmp_path: Path):
        p = tmp_path / "cssci.txt"
        p.write_text(
            "来源篇名\t来源作者\t期刊名称\t年份\t机构\t基金\n"
            "标题\t甲\t期刊\t2024\t清华大学；北京大学\t基金A；基金B\n",
            encoding="utf-8",
        )
        r = CssciParser().parse(p)[0]
        assert r.institutions == [
            Institution(name="清华大学"),
            Institution(name="北京大学"),
        ]
        assert r.funding == ["基金A", "基金B"]
