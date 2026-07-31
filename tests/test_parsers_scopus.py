"""Edge-case coverage tests for the Scopus parser."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from citationer.models.record import Author, DocType
from citationer.parsers.scopus import ScopusParser


class TestScopusParserSource:
    def test_source_name(self):
        assert ScopusParser().source_name == "Scopus"


class TestScopusDetect:
    def test_detect_csv(self, tmp_path: Path):
        p = tmp_path / "scopus.csv"
        p.write_text(
            "Authors,Title,Year,Source title,DOI\n"
            "Smith, J.,ML paper,2024,Journal of AI,10.1000/x\n",
            encoding="utf-8-sig",
        )
        assert ScopusParser().detect(p) is True

    def test_detect_rejects_non_csv_xlsx(self, tmp_path: Path):
        p = tmp_path / "scopus.pdf"
        p.write_text("not scopus", encoding="utf-8")
        assert ScopusParser().detect(p) is False

    def test_detect_rejects_only_two_markers(self, tmp_path: Path):
        p = tmp_path / "scopus.csv"
        p.write_text("Authors,Title\nSmith,Paper\n", encoding="utf-8-sig")
        assert ScopusParser().detect(p) is False

    def test_detect_xlsx(self, tmp_path: Path):
        p = tmp_path / "scopus.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Authors", "Title", "Year", "Source title", "DOI"])
        wb.save(p)
        assert ScopusParser().detect(p) is True

    def test_detect_empty_csv_returns_false(self, tmp_path: Path):
        p = tmp_path / "empty.csv"
        p.write_text("", encoding="utf-8-sig")
        assert ScopusParser().detect(p) is False

    def test_detect_corrupt_xlsx_returns_false(self, tmp_path: Path):
        p = tmp_path / "corrupt.xlsx"
        p.write_bytes(b"not an xlsx file")
        assert ScopusParser().detect(p) is False


class TestScopusParseCsv:
    def test_parse_full_record(self, tmp_path: Path):
        p = tmp_path / "scopus.csv"
        p.write_text(
            "Authors,Author(s) ID,Title,Year,Source title,Volume,Issue,Pages,"
            "Article Number,DOI,Abstract,Author Keywords,Index Keywords,"
            "Language of Original Document,Document Type,Cited by,References,"
            "ISSN,Funding Details,Affiliations\n"
            '"Smith, J.;Jones, M.",12345,Test paper,2024,Journal of X,10,2,100-120,9,'
            '10.1000/t,Abstract text.,keyword1;keyword2,idx1;idx2,English,Review,5,'
            '"Ref 1;Ref 2",1234-5678,"Grant A;Grant B",'
            '"MIT, Cambridge, MA USA; Harvard Univ, Cambridge, MA USA"\n',
            encoding="utf-8-sig",
        )
        records = ScopusParser().parse(p)
        assert len(records) == 1
        r = records[0]
        assert r.title == "Test paper"
        assert r.year == 2024
        assert r.journal == "Journal of X"
        assert r.volume == "10"
        assert r.issue == "2"
        assert r.pages == "100-120"
        assert r.doi == "10.1000/t"
        assert r.issn == "1234-5678"
        assert r.abstract == "Abstract text."
        assert set(r.keywords) == {"keyword1", "keyword2", "idx1", "idx2"}
        assert r.doc_type == DocType.REVIEW
        assert r.citation_count == 5
        assert r.references == ["Ref 1", "Ref 2"]
        assert r.funding == ["Grant A", "Grant B"]
        assert r.source_database == "Scopus"
        assert r.language == "English"
        assert len(r.authors) == 2
        assert r.authors[0].surname == "Smith"
        assert r.authors[0].given_name == "J."
        # Affiliations
        assert len(r.institutions) == 2
        names = {i.name for i in r.institutions}
        assert "MIT" in names
        assert "Harvard Univ" in names

    def test_parse_empty_file_returns_empty(self, tmp_path: Path):
        p = tmp_path / "empty.csv"
        p.write_text("", encoding="utf-8-sig")
        assert ScopusParser().parse(p) == []

    def test_parse_blank_rows_skipped(self, tmp_path: Path):
        p = tmp_path / "scopus.csv"
        p.write_text(
            "Authors,Title,Year,Source title,DOI\n"
            "\n"
            "Smith,Paper,2024,Journal,10.1000/x\n"
            ",,,,\n",
            encoding="utf-8-sig",
        )
        records = ScopusParser().parse(p)
        assert len(records) == 1
        assert records[0].title == "Paper"

    def test_parse_missing_columns_defaults(self, tmp_path: Path):
        p = tmp_path / "scopus.csv"
        p.write_text("Title,Year\nMinimal,2023\n", encoding="utf-8-sig")
        records = ScopusParser().parse(p)
        assert len(records) == 1
        r = records[0]
        assert r.title == "Minimal"
        assert r.year == 2023
        assert r.authors == []
        assert r.keywords == []
        assert r.doi is None

    def test_parse_invalid_year_returns_none(self, tmp_path: Path):
        p = tmp_path / "scopus.csv"
        p.write_text(
            "Authors,Title,Year,Source title,DOI\n"
            "Smith,Paper,no year,Journal,10.1000/x\n",
            encoding="utf-8-sig",
        )
        records = ScopusParser().parse(p)
        assert records[0].year is None

    def test_parse_author_without_comma(self, tmp_path: Path):
        p = tmp_path / "scopus.csv"
        p.write_text(
            "Authors,Title,Year,Source title,DOI\n"
            "Smith J,Paper,2024,Journal,10.1000/x\n",
            encoding="utf-8-sig",
        )
        r = ScopusParser().parse(p)[0]
        assert r.authors[0].full_name == "Smith J"
        assert r.authors[0].surname == "Smith J"
        assert r.authors[0].given_name is None

    def test_parse_doc_type_article(self, tmp_path: Path):
        p = tmp_path / "scopus.csv"
        p.write_text(
            "Authors,Title,Year,Source title,DOI,Document Type\n"
            "Smith,Paper,2024,Journal,10.1000/x,Article\n",
            encoding="utf-8-sig",
        )
        assert ScopusParser().parse(p)[0].doc_type == DocType.ARTICLE

    def test_parse_doc_type_conference_paper(self, tmp_path: Path):
        p = tmp_path / "scopus.csv"
        p.write_text(
            "Authors,Title,Year,Source title,DOI,Document Type\n"
            "Smith,Paper,2024,Journal,10.1000/x,Conference Paper\n",
            encoding="utf-8-sig",
        )
        assert ScopusParser().parse(p)[0].doc_type == DocType.CONFERENCE

    def test_parse_doc_type_unknown(self, tmp_path: Path):
        p = tmp_path / "scopus.csv"
        p.write_text(
            "Authors,Title,Year,Source title,DOI,Document Type\n"
            "Smith,Paper,2024,Journal,10.1000/x,Dataset\n",
            encoding="utf-8-sig",
        )
        assert ScopusParser().parse(p)[0].doc_type == DocType.UNKNOWN

    def test_parse_invalid_citation_count(self, tmp_path: Path):
        p = tmp_path / "scopus.csv"
        p.write_text(
            "Authors,Title,Year,Source title,DOI,Cited by\n"
            "Smith,Paper,2024,Journal,10.1000/x,N/A\n",
            encoding="utf-8-sig",
        )
        assert ScopusParser().parse(p)[0].citation_count is None

    def test_parse_pages_fallback_to_art_no(self, tmp_path: Path):
        p = tmp_path / "scopus.csv"
        p.write_text(
            "Authors,Title,Year,Source title,DOI,Article Number\n"
            "Smith,Paper,2024,Journal,10.1000/x,789\n",
            encoding="utf-8-sig",
        )
        assert ScopusParser().parse(p)[0].pages == "789"


class TestScopusParseXlsx:
    def _save_xlsx(self, path: Path, rows: list[list[str]]) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        wb.save(path)

    def test_parse_xlsx_full_record(self, tmp_path: Path):
        p = tmp_path / "scopus.xlsx"
        self._save_xlsx(
            p,
            [
                ["Authors", "Title", "Year", "Source title", "DOI", "Document Type"],
                ["Smith, J.", "XLSX Paper", "2023", "Journal Y", "10.1000/y", "Book Chapter"],
            ],
        )
        records = ScopusParser().parse(p)
        assert len(records) == 1
        r = records[0]
        assert r.title == "XLSX Paper"
        assert r.year == 2023
        assert r.doi == "10.1000/y"
        assert r.doc_type == DocType.BOOK_CHAPTER
        assert r.authors[0].surname == "Smith"

    def test_parse_xlsx_empty_workbook_returns_empty(self, tmp_path: Path):
        p = tmp_path / "empty.xlsx"
        openpyxl.Workbook().save(p)
        assert ScopusParser().parse(p) == []

    def test_parse_xlsx_blank_rows_skipped(self, tmp_path: Path):
        p = tmp_path / "scopus.xlsx"
        self._save_xlsx(
            p,
            [
                ["Authors", "Title", "Year", "Source title", "DOI"],
                ["Smith, J.", "First", "2024", "Journal", "10.1000/a"],
                [None, None, None, None, None],
                ["Jones, M.", "Second", "2023", "Journal", "10.1000/b"],
            ],
        )
        records = ScopusParser().parse(p)
        assert len(records) == 2
        assert records[0].title == "First"
        assert records[1].title == "Second"

    def test_parse_xlsx_short_row(self, tmp_path: Path):
        p = tmp_path / "scopus.xlsx"
        self._save_xlsx(
            p,
            [
                ["Authors", "Title", "Year", "Source title", "DOI"],
                ["Smith, J.", "Short"],
            ],
        )
        records = ScopusParser().parse(p)
        assert len(records) == 1
        assert records[0].title == "Short"
        assert records[0].year is None


class TestScopusParseHelpers:
    @pytest.mark.parametrize(
        "raw,expected",
        [
            ("", []),
            ("  ;  ", []),
            (
                "Smith, J.; Jones, M.",
                [
                    Author(full_name="Smith, J.", surname="Smith", given_name="J.", order=1),
                    Author(full_name="Jones, M.", surname="Jones", given_name="M.", order=2),
                ],
            ),
        ],
    )
    def test_parse_authors(self, raw: str, expected: list[Author]):
        assert ScopusParser._parse_authors(raw) == expected

    def test_parse_doc_type_editorial(self):
        assert ScopusParser._parse_doc_type("Editorial") == DocType.OTHER

    def test_parse_doc_type_erratum(self):
        assert ScopusParser._parse_doc_type("Erratum") == DocType.OTHER
