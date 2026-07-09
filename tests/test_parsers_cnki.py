"""Tests for CNKI (Chinese) Excel parser.

Covers:
- Header detection
- Full record parsing (title, authors, institutions, keywords, abstract, etc.)
- Author parsing (Chinese names with ";" / "；" separators)
- Institution parsing
- Keyword parsing (";;" and ";" fallbacks)
- Doc type mapping (Chinese & English variants)
- Year extraction from various date formats
- Funding field
- Edge cases: empty rows, missing columns, malformed data
"""

from __future__ import annotations

import openpyxl
import pytest
from openpyxl import Workbook

from citationer.models.record import DocType
from citationer.parsers.cnki import CnkiExcelParser


def _make_cnki_xlsx(headers: list[str], rows: list[list], path) -> None:
    """Helper to create a CNKI-formatted xlsx file."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


# Standard CNKI column headers (Chinese)
CNKI_HEADERS = [
    "题名", "作者", "机构", "来源", "发表时间", "关键词",
    "摘要", "DOI", "基金", "卷", "期", "页", "文献类型", "被引",
]


class TestCnkiDetect:
    def test_detect_valid_cnki_xlsx(self, tmp_path):
        f = tmp_path / "cnki.xlsx"
        _make_cnki_xlsx(CNKI_HEADERS, [], f)
        assert CnkiExcelParser().detect(f) is True

    def test_reject_non_xlsx(self, tmp_path):
        f = tmp_path / "test.txt"
        f.write_text("题名 作者 来源")
        assert CnkiExcelParser().detect(f) is False

    def test_reject_wrong_headers(self, tmp_path):
        f = tmp_path / "wrong.xlsx"
        wb = Workbook()
        wb.active.append(["name", "date", "value"])  # not CNKI markers
        wb.save(f)
        assert CnkiExcelParser().detect(f) is False

    def test_partial_headers_min_3_markers(self, tmp_path):
        """Need at least 3 CNKI markers (e.g., 题名 + 作者 + 来源)."""
        f = tmp_path / "partial.xlsx"
        wb = Workbook()
        wb.active.append(["题名", "作者", "来源", "extra"])  # 3 markers
        wb.save(f)
        assert CnkiExcelParser().detect(f) is True

    def test_only_2_markers_rejected(self, tmp_path):
        f = tmp_path / "few.xlsx"
        wb = Workbook()
        wb.active.append(["题名", "作者", "extra"])
        wb.save(f)
        assert CnkiExcelParser().detect(f) is False

    def test_corrupt_xlsx_returns_false(self, tmp_path):
        f = tmp_path / "bad.xlsx"
        f.write_text("not a real xlsx file")
        assert CnkiExcelParser().detect(f) is False

    def test_source_name(self):
        assert CnkiExcelParser().source_name == "CNKI"


class TestCnkiParse:
    def test_parse_empty_xlsx(self, tmp_path):
        """Empty xlsx (header only) returns empty list."""
        f = tmp_path / "empty.xlsx"
        _make_cnki_xlsx(CNKI_HEADERS, [], f)
        assert CnkiExcelParser().parse(f) == []

    def test_parse_single_full_record(self, tmp_path):
        f = tmp_path / "test.xlsx"
        _make_cnki_xlsx(
            CNKI_HEADERS,
            [[
                "深度学习在图像识别中的应用",
                "张伟;李娜;王强",
                "清华大学;北京大学",
                "计算机学报",
                "2023-05-15",
                "深度学习;;图像识别",
                "本文研究了深度学习在图像识别中的应用。",
                "10.1000/test.2023",
                "国家自然科学基金;国家重点研发计划",
                "46",
                "5",
                "1-10",
                "期刊论文",
                "25",
            ]],
            f,
        )
        records = CnkiExcelParser().parse(f)
        assert len(records) == 1
        r = records[0]
        assert r.title == "深度学习在图像识别中的应用"
        assert r.year == 2023
        assert r.journal == "计算机学报"
        assert r.doi == "10.1000/test.2023"
        assert r.volume == "46"
        assert r.issue == "5"
        assert r.pages == "1-10"
        assert r.citation_count == 25
        assert r.language == "zh"
        assert r.source_database == "CNKI"
        assert r.doc_type == DocType.ARTICLE

    def test_authors_parsed(self, tmp_path):
        f = tmp_path / "test.xlsx"
        _make_cnki_xlsx(
            CNKI_HEADERS,
            [["论文", "张伟;李娜;王强", "", "期刊", "2024", "", "", "", "", "", "", "", "", ""]],
            f,
        )
        r = CnkiExcelParser().parse(f)[0]
        assert len(r.authors) == 3
        assert r.authors[0].full_name == "张伟"
        assert r.authors[0].order == 1
        assert r.authors[2].full_name == "王强"
        assert r.authors[2].order == 3

    def test_authors_chinese_semicolon_separator(self, tmp_path):
        """CNKI uses '；' (full-width semicolon) in some exports."""
        f = tmp_path / "test.xlsx"
        _make_cnki_xlsx(
            CNKI_HEADERS,
            [["论文", "张伟；李娜", "", "期刊", "2024", "", "", "", "", "", "", "", "", ""]],
            f,
        )
        r = CnkiExcelParser().parse(f)[0]
        assert len(r.authors) == 2
        assert r.authors[0].full_name == "张伟"
        assert r.authors[1].full_name == "李娜"

    def test_institutions_parsed(self, tmp_path):
        f = tmp_path / "test.xlsx"
        _make_cnki_xlsx(
            CNKI_HEADERS,
            [["论文", "", "清华大学;北京大学;中科院", "期刊", "2024", "", "", "", "", "", "", "", "", ""]],
            f,
        )
        r = CnkiExcelParser().parse(f)[0]
        assert len(r.institutions) == 3
        assert r.institutions[0].name == "清华大学"

    def test_keywords_double_semicolon(self, tmp_path):
        """CNKI uses ';;' to separate keywords."""
        f = tmp_path / "test.xlsx"
        _make_cnki_xlsx(
            CNKI_HEADERS,
            [["论文", "", "", "期刊", "2024", "深度学习;;图像识别;;神经网络", "", "", "", "", "", "", "", ""]],
            f,
        )
        r = CnkiExcelParser().parse(f)[0]
        assert r.keywords == ["深度学习", "图像识别", "神经网络"]

    def test_keywords_single_semicolon_fallback(self, tmp_path):
        f = tmp_path / "test.xlsx"
        _make_cnki_xlsx(
            CNKI_HEADERS,
            [["论文", "", "", "期刊", "2024", "深度学习;图像识别", "", "", "", "", "", "", "", ""]],
            f,
        )
        r = CnkiExcelParser().parse(f)[0]
        assert r.keywords == ["深度学习", "图像识别"]

    def test_funding_parsed(self, tmp_path):
        f = tmp_path / "test.xlsx"
        _make_cnki_xlsx(
            CNKI_HEADERS,
            [["论文", "", "", "期刊", "2024", "", "", "", "国家自然科学基金;国家重点研发计划", "", "", "", "", ""]],
            f,
        )
        r = CnkiExcelParser().parse(f)[0]
        assert r.funding == ["国家自然科学基金", "国家重点研发计划"]

    def test_year_from_date(self, tmp_path):
        """Year extracted from various date formats."""
        f = tmp_path / "test.xlsx"
        for date, expected_year in [
            ("2024-05-15", 2024),
            ("2023/12/01", 2023),
            ("2022", 2022),
            ("2020-06", 2020),
        ]:
            _make_cnki_xlsx(
                CNKI_HEADERS,
                [["论文", "", "", "期刊", date, "", "", "", "", "", "", "", "", ""]],
                f,
            )
            r = CnkiExcelParser().parse(f)[0]
            assert r.year == expected_year, f"Failed for {date}"

    def test_year_invalid_returns_none(self, tmp_path):
        f = tmp_path / "test.xlsx"
        _make_cnki_xlsx(
            CNKI_HEADERS,
            [["论文", "", "", "期刊", "not a date", "", "", "", "", "", "", "", "", ""]],
            f,
        )
        r = CnkiExcelParser().parse(f)[0]
        assert r.year is None

    def test_year_empty(self, tmp_path):
        f = tmp_path / "test.xlsx"
        _make_cnki_xlsx(
            CNKI_HEADERS,
            [["论文", "", "", "期刊", "", "", "", "", "", "", "", "", "", ""]],
            f,
        )
        r = CnkiExcelParser().parse(f)[0]
        assert r.year is None

    def test_doc_type_article(self, tmp_path):
        f = tmp_path / "test.xlsx"
        _make_cnki_xlsx(
            CNKI_HEADERS,
            [["论文", "", "", "期刊", "2024", "", "", "", "", "", "", "", "期刊论文", ""]],
            f,
        )
        r = CnkiExcelParser().parse(f)[0]
        assert r.doc_type == DocType.ARTICLE

    def test_doc_type_conference(self, tmp_path):
        f = tmp_path / "test.xlsx"
        _make_cnki_xlsx(
            CNKI_HEADERS,
            [["论文", "", "", "期刊", "2024", "", "", "", "", "", "", "", "会议论文", ""]],
            f,
        )
        r = CnkiExcelParser().parse(f)[0]
        assert r.doc_type == DocType.CONFERENCE

    def test_doc_type_thesis(self, tmp_path):
        f = tmp_path / "test.xlsx"
        _make_cnki_xlsx(
            CNKI_HEADERS,
            [["论文", "", "", "期刊", "2024", "", "", "", "", "", "", "", "博士学位论文", ""]],
            f,
        )
        r = CnkiExcelParser().parse(f)[0]
        assert r.doc_type == DocType.THESIS

    def test_doc_type_patent(self, tmp_path):
        f = tmp_path / "test.xlsx"
        _make_cnki_xlsx(
            CNKI_HEADERS,
            [["论文", "", "", "期刊", "2024", "", "", "", "", "", "", "", "专利", ""]],
            f,
        )
        r = CnkiExcelParser().parse(f)[0]
        assert r.doc_type == DocType.PATENT

    def test_doc_type_book(self, tmp_path):
        f = tmp_path / "test.xlsx"
        _make_cnki_xlsx(
            CNKI_HEADERS,
            [["论文", "", "", "期刊", "2024", "", "", "", "", "", "", "", "图书", ""]],
            f,
        )
        r = CnkiExcelParser().parse(f)[0]
        assert r.doc_type == DocType.BOOK

    def test_doc_type_unknown_falls_to_other(self, tmp_path):
        f = tmp_path / "test.xlsx"
        _make_cnki_xlsx(
            CNKI_HEADERS,
            [["论文", "", "", "期刊", "2024", "", "", "", "", "", "", "", "其他类型", ""]],
            f,
        )
        r = CnkiExcelParser().parse(f)[0]
        assert r.doc_type == DocType.OTHER

    def test_citation_count_invalid(self, tmp_path):
        f = tmp_path / "test.xlsx"
        _make_cnki_xlsx(
            CNKI_HEADERS,
            [["论文", "", "", "期刊", "2024", "", "", "", "", "", "", "", "", "not a number"]],
            f,
        )
        r = CnkiExcelParser().parse(f)[0]
        assert r.citation_count is None

    def test_empty_rows_skipped(self, tmp_path):
        """All-empty data rows should be skipped."""
        f = tmp_path / "test.xlsx"
        _make_cnki_xlsx(
            CNKI_HEADERS,
            [
                ["论文1", "张伟", "", "期刊", "2024", "", "", "", "", "", "", "", "", ""],
                ["", "", "", "", "", "", "", "", "", "", "", "", "", ""],  # empty
                ["论文2", "李娜", "", "期刊", "2024", "", "", "", "", "", "", "", "", ""],
            ],
            f,
        )
        records = CnkiExcelParser().parse(f)
        assert len(records) == 2

    def test_source_file_propagated(self, tmp_path):
        f = tmp_path / "my_paper_collection.xlsx"
        _make_cnki_xlsx(
            CNKI_HEADERS,
            [["论文", "", "", "期刊", "2024", "", "", "", "", "", "", "", "", ""]],
            f,
        )
        r = CnkiExcelParser().parse(f)[0]
        assert r.source_file == "my_paper_collection.xlsx"

    def test_alternate_column_units(self, tmp_path):
        """CNKI sometimes uses '单位' instead of '机构' for institutions."""
        f = tmp_path / "test.xlsx"
        headers = ["题名", "作者", "单位", "来源", "发表时间", "关键词",
                   "摘要", "DOI", "基金", "卷", "期", "页", "文献类型", "被引"]
        _make_cnki_xlsx(
            headers,
            [["论文", "", "清华大学", "期刊", "2024", "", "", "", "", "", "", "", "", ""]],
            f,
        )
        r = CnkiExcelParser().parse(f)[0]
        assert len(r.institutions) == 1
        assert r.institutions[0].name == "清华大学"

    def test_alternate_column_year(self, tmp_path):
        """CNKI sometimes uses '年' instead of '发表时间'."""
        f = tmp_path / "test.xlsx"
        headers = ["题名", "作者", "机构", "来源", "年", "关键词",
                   "摘要", "DOI", "基金", "卷", "期", "页", "文献类型", "被引"]
        _make_cnki_xlsx(
            headers,
            [["论文", "", "", "期刊", "2023", "", "", "", "", "", "", "", "", ""]],
            f,
        )
        r = CnkiExcelParser().parse(f)[0]
        assert r.year == 2023
