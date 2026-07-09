"""Tests for Phase 4 additions: new parsers, interactive wizard, pipeline runner."""


from citationer.models.record import Author
from citationer.parsers.bibtex import BibTeXParser
from citationer.parsers.cssci import CssciParser
from citationer.parsers.pubmed import PubMedParser
from citationer.parsers.ris import RISParser
from citationer.parsers.scopus import ScopusParser

# ===========================================================================
# BibTeX parser
# ===========================================================================


class TestBibTeXParser:
    def setup_method(self):
        self.parser = BibTeXParser()

    def test_detect_bib_file(self, tmp_path):
        p = tmp_path / "refs.bib"
        p.write_text("@article{key, title={Test}}", encoding="utf-8")
        assert self.parser.detect(p) is True

    def test_detect_non_bib_file(self, tmp_path):
        p = tmp_path / "refs.txt"
        p.write_text("not bibtex", encoding="utf-8")
        assert self.parser.detect(p) is False

    def test_parse_article(self, tmp_path):
        p = tmp_path / "refs.bib"
        p.write_text(
            "@article{key1,\n"
            "  title = {ML in Healthcare},\n"
            "  author = {Smith, John and Jones, Mary},\n"
            "  journal = {Journal of AI},\n"
            "  year = {2024},\n"
            "  doi = {10.1000/test}\n"
            "}",
            encoding="utf-8",
        )
        records = self.parser.parse(p)
        assert len(records) == 1
        r = records[0]
        assert r.title == "ML in Healthcare"
        assert r.year == 2024
        assert r.journal == "Journal of AI"
        assert r.doi == "10.1000/test"
        assert r.doc_type.value == "article"
        assert len(r.authors) == 2
        assert r.authors[0].surname == "Smith"
        assert r.authors[0].order == 1

    def test_parse_inproceedings(self, tmp_path):
        p = tmp_path / "conf.bib"
        p.write_text(
            "@inproceedings{key1,\n"
            "  title = {NLP for Biomed},\n"
            "  author = {Brown, Alice},\n"
            "  booktitle = {Proc. NLP Conf.},\n"
            "  year = {2023}\n"
            "}",
            encoding="utf-8",
        )
        records = self.parser.parse(p)
        assert records[0].doc_type.value == "book_chapter"

    def test_parse_multiple_entries(self, tmp_path):
        p = tmp_path / "multi.bib"
        p.write_text(
            "@article{a, title={A}, year={2020}}\n"
            "@article{b, title={B}, year={2021}}\n"
            "@article{c, title={C}, year={2022}}\n",
            encoding="utf-8",
        )
        records = self.parser.parse(p)
        assert len(records) == 3

    def test_parse_empty_file(self, tmp_path):
        p = tmp_path / "empty.bib"
        p.write_text("", encoding="utf-8")
        assert self.parser.parse(p) == []

    def test_parse_entry_without_title(self, tmp_path):
        p = tmp_path / "no_title.bib"
        p.write_text(
            "@article{x, author = {Doe, J}, year = {2020}}",
            encoding="utf-8",
        )
        # Entries without title are skipped
        assert self.parser.parse(p) == []


# ===========================================================================
# RIS parser
# ===========================================================================


class TestRISParser:
    def setup_method(self):
        self.parser = RISParser()

    def test_detect_ris_file(self, tmp_path):
        p = tmp_path / "refs.ris"
        p.write_text("TY  - JOUR\nTI  - Test\nER  - \n", encoding="utf-8")
        assert self.parser.detect(p) is True

    def test_detect_non_ris(self, tmp_path):
        p = tmp_path / "data.csv"
        p.write_text("a,b,c", encoding="utf-8")
        assert self.parser.detect(p) is False

    def test_parse_single_record(self, tmp_path):
        p = tmp_path / "refs.ris"
        p.write_text(
            "TY  - JOUR\n"
            "TI  - Sample Paper\n"
            "AU  - Smith, John\n"
            "AU  - Jones, Mary\n"
            "PY  - 2024\n"
            "JO  - Journal of Testing\n"
            "VL  - 10\n"
            "IS  - 2\n"
            "SP  - 100\n"
            "EP  - 120\n"
            "DO  - 10.1000/test\n"
            "AB  - Abstract text.\n"
            "KW  - keyword1\n"
            "KW  - keyword2\n"
            "LA  - English\n"
            "ER  - \n",
            encoding="utf-8",
        )
        records = self.parser.parse(p)
        assert len(records) == 1
        r = records[0]
        assert r.title == "Sample Paper"
        assert r.year == 2024
        assert r.journal == "Journal of Testing"
        assert r.doi == "10.1000/test"
        assert r.doc_type.value == "article"
        assert len(r.authors) == 2
        assert "keyword1" in r.keywords
        assert r.language == "eng"

    def test_parse_multiple_records(self, tmp_path):
        p = tmp_path / "refs.ris"
        p.write_text(
            "TY  - JOUR\nTI  - First\nAU  - A, B\nPY  - 2020\nER  - \n"
            "TY  - JOUR\nTI  - Second\nAU  - C, D\nPY  - 2021\nER  - \n",
            encoding="utf-8",
        )
        records = self.parser.parse(p)
        assert len(records) == 2
        assert records[0].title == "First"
        assert records[1].title == "Second"

    def test_parse_continuation_lines(self, tmp_path):
        p = tmp_path / "refs.ris"
        p.write_text(
            "TY  - JOUR\n"
            "TI  - Test paper\n"
            "AB  - This is a long abstract that spans\n"
            "    multiple lines.\n"
            "ER  - \n",
            encoding="utf-8",
        )
        records = self.parser.parse(p)
        assert records is not None and len(records) == 1
        assert "multiple lines" in records[0].abstract

    def test_no_er_uses_ty_boundary(self, tmp_path):
        """When ER is missing, split records on TY boundary."""
        p = tmp_path / "no_er.ris"
        p.write_text(
            "TY  - JOUR\nTI  - First\nAU  - A\nPY  - 2020\n"
            "TY  - JOUR\nTI  - Second\nAU  - B\nPY  - 2021\n",
            encoding="utf-8",
        )
        records = self.parser.parse(p)
        assert len(records) == 2

    def test_parse_pages_from_sp_ep(self, tmp_path):
        p = tmp_path / "pages.ris"
        p.write_text(
            "TY  - JOUR\n"
            "TI  - P\n"
            "SP  - 100\n"
            "EP  - 120\n"
            "ER  - \n",
            encoding="utf-8",
        )
        records = self.parser.parse(p)
        assert records[0].pages == "100-120"

    def test_doc_type_mapping(self, tmp_path):
        for ris_type, expected in [
            ("JOUR", "article"),
            ("BOOK", "book"),
            ("CHAP", "book_chapter"),
            ("CONF", "conference"),
        ]:
            p = tmp_path / "t.ris"
            p.write_text(
                f"TY  - {ris_type}\nTI  - X\nER  - \n",
                encoding="utf-8",
            )
            records = self.parser.parse(p)
            assert records[0].doc_type.value == expected, ris_type


# ===========================================================================
# PubMed parser
# ===========================================================================


class TestPubMedParser:
    def setup_method(self):
        self.parser = PubMedParser()

    def test_detect_xml_file(self, tmp_path):
        p = tmp_path / "pubmed.xml"
        p.write_text(
            "<?xml version='1.0'?><PubmedArticle>"
            "<MedlineCitation><Article><ArticleTitle>X</ArticleTitle>"
            "</Article></MedlineCitation></PubmedArticle>",
            encoding="utf-8",
        )
        assert self.parser.detect(p) is True

    def test_detect_non_pubmed(self, tmp_path):
        p = tmp_path / "data.txt"
        p.write_text("not pubmed", encoding="utf-8")
        assert self.parser.detect(p) is False


# ===========================================================================
# CSSCI parser
# ===========================================================================


class TestCssciParser:
    def setup_method(self):
        self.parser = CssciParser()

    def test_detect_xlsx(self, tmp_path):
        import openpyxl
        p = tmp_path / "cssci.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["来源篇名", "来源作者", "期刊名称", "年份"])
        wb.save(p)
        assert self.parser.detect(p) is True

    def test_detect_non_cssci(self, tmp_path):
        import openpyxl
        p = tmp_path / "other.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["random", "columns"])
        wb.save(p)
        assert self.parser.detect(p) is False

    def test_parse_xlsx(self, tmp_path):
        import openpyxl
        p = tmp_path / "cssci.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["来源篇名", "来源作者", "期刊名称", "年份", "DOI", "摘要", "关键词"])
        ws.append(["中文标题A", "作者甲;作者乙", "CSSCI期刊", "2024",
                    "10.1000/cssci1", "这是中文摘要。", "关键词1;关键词2"])
        wb.save(p)
        records = self.parser.parse(p)
        assert len(records) == 1
        r = records[0]
        assert r.title == "中文标题A"
        assert r.year == 2024
        assert r.language == "zh"
        assert r.source_database == "CSSCI"
        assert len(r.authors) == 2
        assert "关键词1" in r.keywords


# ===========================================================================
# Scopus parser
# ===========================================================================


class TestScopusParser:
    def setup_method(self):
        self.parser = ScopusParser()

    def test_detect_csv(self, tmp_path):
        p = tmp_path / "scopus.csv"
        p.write_text(
            "Authors,Title,Year,Source title,DOI\n"
            "Smith, J.,ML paper,2024,Journal of AI,10.1000/x\n",
            encoding="utf-8",
        )
        assert self.parser.detect(p) is True

    def test_detect_non_scopus(self, tmp_path):
        p = tmp_path / "data.csv"
        p.write_text("a,b,c\n1,2,3\n", encoding="utf-8")
        assert self.parser.detect(p) is False

    def test_parse_csv(self, tmp_path):
        p = tmp_path / "s.csv"
        p.write_text(
            "Authors,Title,Year,Source title,DOI,Abstract,Author Keywords,"
            "Document Type,Cited by\n"
            '"Smith, J.;Jones, M.",Test paper,2024,Journal of X,'
            '10.1000/t,"Abstract.","keyword1;keyword2",Article,5\n',
            encoding="utf-8",
        )
        records = self.parser.parse(p)
        assert len(records) == 1
        r = records[0]
        assert r.title == "Test paper"
        assert r.year == 2024
        assert r.citation_count == 5
        assert "keyword1" in r.keywords


# ===========================================================================
# Pipeline runner
# ===========================================================================


class TestPipelineRunner:
    def test_run_stats_overview(self, tmp_path):
        from typer.testing import CliRunner

        from citationer.cli.run_cmd import app
        from citationer.models.record import Record

        # Need a populated database first
        from citationer.utils.database import CitationDatabase
        db_path = tmp_path / "test.db"
        db = CitationDatabase(db_path)
        db.initialize()
        rec = Record(
            title="Test", authors=[Author(full_name="X")], year=2024,
            source_database="test",
        )
        from citationer.utils.serialization import record_to_db_serializable
        db.insert_record(**record_to_db_serializable(rec))
        db.close()

        cfg = tmp_path / "pipeline.yaml"
        cfg.write_text(
            "name: test\n"
            "output_dir: " + str(tmp_path / "out") + "\n"
            "steps:\n"
            "  - name: overview\n"
            "    action: stats\n"
            "    type: overview\n"
            "    output: overview.txt\n",
            encoding="utf-8",
        )

        # Mock DB path
        from citationer.utils import config as cfg_module
        original = cfg_module.get_db_path
        cfg_module.get_db_path = lambda: db_path
        try:
            runner = CliRunner()
            result = runner.invoke(app, [str(cfg)])
            assert result.exit_code == 0
            assert (tmp_path / "out" / "overview.txt").exists()
        finally:
            cfg_module.get_db_path = original

    def test_pipeline_validation(self, tmp_path):
        import pytest as _pytest

        from citationer.cli.run_cmd import _execute_step
        with _pytest.raises(ValueError, match="未知 action"):
            _execute_step("bogus", {}, [], {}, tmp_path)
