"""CSSCI (中文社会科学引文索引) export parser."""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from citationer.models.record import Author, DocType, Institution, Record
from citationer.parsers.base import BaseParser


class CssciParser(BaseParser):
    """Parser for CSSCI (Chinese Social Sciences Citation Index) exports.

    CSSCI exports typically contain Chinese column headers such as
    ``来源篇名``, ``来源作者``, ``期刊名称``, ``年份``, etc.
    """

    CSSCI_MARKERS: list[str] = [
        "来源篇名",
        "来源作者",
        "期刊名称",
        "年份",
    ]

    COLUMN_MAP: dict[str, str] = {
        "来源篇名": "title",
        "英文篇名": "title_en",
        "来源作者": "authors_raw",
        "作者": "authors_raw",
        "期刊名称": "journal",
        "英文刊名": "journal_en",
        "年份": "year_raw",
        "年": "year_raw",
        "卷": "volume",
        "期": "issue",
        "页码": "pages",
        "关键词": "keywords_raw",
        "摘要": "abstract",
        "机构": "institutions_raw",
        "作者机构": "institutions_raw",
        "基金": "funding_raw",
        "学科分类": "subject",
        "被引频次": "citation_count",
        "DOI": "doi",
    }

    def __init__(self, encoding: str | None = None) -> None:
        self._encoding = encoding

    @staticmethod
    def _detect_encoding(filepath: Path) -> str:
        """Detect text file encoding by trying common Chinese encodings."""
        for enc in ("utf-8", "utf-8-sig", "gbk", "gb2312", "gb18030"):
            try:
                with open(filepath, encoding=enc) as f:
                    f.read()
                return enc
            except (UnicodeDecodeError, UnicodeError):
                continue
        return "utf-8"

    @property
    def source_name(self) -> str:
        return "CSSCI"

    def detect(self, filepath: Path) -> bool:
        suffix = filepath.suffix.lower()
        if suffix not in (".xlsx", ".xls", ".txt", ".csv"):
            return False

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                wb.close()
                return False
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            wb.close()
            if row is None:
                return False
            headers = [str(c).strip() if c else "" for c in row]
            header_text = " ".join(headers)
            match_count = sum(1 for m in self.CSSCI_MARKERS if m in header_text)
            return match_count >= 2
        except Exception:
            pass

        # Try text file
        if suffix in (".txt", ".csv"):
            encoding = self._encoding or self._detect_encoding(filepath)
            try:
                with open(filepath, encoding=encoding, errors="ignore") as f:
                    first = f.readline()
                return any(m in first for m in self.CSSCI_MARKERS)
            except Exception:
                pass

        return False

    def parse(self, filepath: Path) -> list[Record]:
        suffix = filepath.suffix.lower()
        if suffix in (".xlsx", ".xls"):
            return self._parse_xlsx(filepath)
        encoding = self._encoding or self._detect_encoding(filepath)
        return self._parse_text(filepath, encoding)

    # ------------------------------------------------------------------
    # XLSX
    # ------------------------------------------------------------------

    def _parse_xlsx(self, filepath: Path) -> list[Record]:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            return []

        rows_iter = ws.iter_rows(values_only=True)
        first = next(rows_iter, None)
        if first is None:
            wb.close()
            return []

        headers = [str(c).strip() if c else "" for c in first]
        col_index = self._build_col_index(headers)

        records: list[Record] = []
        for row in rows_iter:
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            records.append(self._parse_row(row, col_index, filepath.name))

        wb.close()
        return records

    # ------------------------------------------------------------------
    # Text/CSV
    # ------------------------------------------------------------------

    def _parse_text(self, filepath: Path, encoding: str) -> list[Record]:
        import csv

        with open(filepath, encoding=encoding, errors="ignore") as f:
            delimiter = "\t" if "\t" in f.readline() else ","
            f.seek(0)
            reader = csv.reader(f, delimiter=delimiter)
            first = next(reader, None)
            if first is None:
                return []

            headers = [str(c).strip() if c else "" for c in first]
            col_index = self._build_col_index(headers)

            records: list[Record] = []
            for row in reader:
                if all(not (c or "").strip() for c in row):
                    continue
                records.append(self._parse_row(tuple(row), col_index, filepath.name))

        return records

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _build_col_index(self, headers: list[str]) -> dict[str, int]:
        index: dict[str, int] = {}
        for i, h in enumerate(headers):
            best_field: str | None = None
            best_len = -1
            for marker, field in self.COLUMN_MAP.items():
                if marker in h and len(marker) > best_len:
                    best_len = len(marker)
                    best_field = field
            if best_field is not None:
                index[best_field] = i
        return index

    def _parse_row(
        self, row: tuple, col_index: dict[str, int], source_file: str
    ) -> Record:
        def val(field: str) -> str:
            idx = col_index.get(field)
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return str(v).strip() if v is not None else ""

        # Year
        year: int | None = None
        y = val("year_raw")
        if y:
            m = re.search(r"(\d{4})", y)
            if m:
                try:
                    year = int(m.group(1))
                except ValueError:
                    pass

        # Authors
        authors: list[Author] = []
        au = val("authors_raw")
        if au:
            for i, name in enumerate(re.split(r"[;；,，]", au)):
                name = name.strip()
                if name:
                    authors.append(Author(full_name=name, order=i + 1))

        # Keywords
        keywords: list[str] = []
        kw = val("keywords_raw")
        if kw:
            for k in re.split(r"[;；,，\s]+", kw):
                k = k.strip()
                if k and len(k) >= 2:
                    keywords.append(k)

        # Institutions
        institutions: list[Institution] = []
        inst = val("institutions_raw")
        if inst:
            for name in re.split(r"[;；]", inst):
                name = name.strip()
                if name:
                    institutions.append(Institution(name=name))

        # Funding
        funding: list[str] | None = None
        fund = val("funding_raw")
        if fund:
            funding = [f.strip() for f in re.split(r"[;；]", fund) if f.strip()]

        # Citation count
        citation_count: int | None = None
        cc = val("citation_count")
        if cc:
            try:
                citation_count = int(cc)
            except ValueError:
                pass

        return Record(
            title=val("title"),
            title_en=val("title_en") or None,
            authors=authors,
            year=year,
            journal=val("journal") or None,
            journal_en=val("journal_en") or None,
            volume=val("volume") or None,
            issue=val("issue") or None,
            pages=val("pages") or None,
            doi=val("doi") or None,
            abstract=val("abstract") or None,
            keywords=keywords,
            language="zh",
            institutions=institutions,
            funding=funding,
            citation_count=citation_count,
            doc_type=DocType.ARTICLE,
            source_database="CSSCI",
            source_file=source_file,
        )
