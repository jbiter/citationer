"""CNKI (知网) Excel export parser."""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from citationer.models.record import Author, DocType, Institution, Record
from citationer.parsers.base import BaseParser


class CnkiExcelParser(BaseParser):
    """Parser for CNKI (中国知网) Excel export files (.xlsx).

    CNKI exports have a characteristic header row with Chinese column names.
    This parser detects CNKI files by checking for those header patterns.
    """

    # Known CNKI column header patterns (normalized, lowercase)
    CNKI_HEADER_MARKERS: list[str] = [
        "题名",
        "作者",
        "来源",
        "发表时间",
        "关键词",
        "摘要",
        "机构",
        "基金",
    ]

    # Mapping from CNKI column names to Record fields
    COLUMN_MAP: dict[str, str] = {
        "题名": "title",
        "作者": "authors_raw",
        "单位": "institutions_raw",
        "机构": "institutions_raw",
        "来源": "journal",
        "发表时间": "year_raw",
        "年": "year_raw",
        "关键词": "keywords_raw",
        "摘要": "abstract",
        "DOI": "doi",
        "基金": "funding_raw",
        "卷": "volume",
        "期": "issue",
        "页": "pages",
        "页码": "pages",
        "文献类型": "doc_type_raw",
        "语言": "language",
        "被引": "citation_count",
    }

    def __init__(self, encoding: str = "utf-8") -> None:
        self._encoding = encoding

    @property
    def source_name(self) -> str:
        return "CNKI"

    def detect(self, filepath: Path) -> bool:
        """Detect CNKI Excel export by checking header row for known markers."""
        if filepath.suffix.lower() not in (".xlsx", ".xls"):
            return False

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                wb.close()
                return False

            # Read the first row (header)
            headers = []
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if row:
                headers = [str(c).strip() if c else "" for c in row]

            wb.close()

            # Check if enough CNKI markers are present
            header_text = " ".join(headers)
            match_count = sum(
                1 for marker in self.CNKI_HEADER_MARKERS if marker in header_text
            )
            return match_count >= 3
        except Exception:
            return False

    def parse(self, filepath: Path) -> list[Record]:
        """Parse a CNKI Excel export into Record objects."""
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            return []

        rows_iter = ws.iter_rows(values_only=True)
        first = next(rows_iter, None)
        if first is None:
            wb.close()
            return []

        header_row = [str(c).strip() if c else "" for c in first]
        col_index = self._build_column_index(header_row)

        records: list[Record] = []
        for row in rows_iter:
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            records.append(self._parse_row(row, col_index, filepath.name))

        wb.close()
        return records

    def _build_column_index(self, headers: list[str]) -> dict[str, int]:
        """Map normalized column names to column indices."""
        index: dict[str, int] = {}
        for i, header in enumerate(headers):
            for cnki_name, field_name in self.COLUMN_MAP.items():
                if cnki_name in header:
                    index[field_name] = i
                    break
        return index

    def _parse_row(
        self,
        row: tuple,
        col_index: dict[str, int],
        source_file: str,
    ) -> Record:
        """Parse a single data row into a Record."""

        def get(field: str) -> str:
            idx = col_index.get(field)
            if idx is None or idx >= len(row):
                return ""
            val = row[idx]
            return str(val).strip() if val is not None else ""

        # Parse year
        year: int | None = None
        year_str = get("year_raw")
        if year_str:
            try:
                # CNKI dates are often like "2024-12-01" or just "2024"
                year_match = re.search(r"(\d{4})", year_str)
                if year_match:
                    year = int(year_match.group(1))
            except ValueError:
                pass

        # Parse authors
        authors = self._parse_authors(get("authors_raw"))

        # Parse institutions
        institutions = self._parse_institutions(get("institutions_raw"))

        # Parse keywords (CNKI uses ";;" or ";" as separator)
        keywords_raw = get("keywords_raw")
        keywords = self._parse_keywords(keywords_raw)

        # Parse doc type
        doc_type = self._parse_doc_type(get("doc_type_raw"))

        # Parse citation count
        citation_count: int | None = None
        cite_str = get("citation_count")
        if cite_str:
            try:
                citation_count = int(cite_str)
            except ValueError:
                pass

        # Parse funding
        funding_raw = get("funding_raw")
        funding: list[str] | None = None
        if funding_raw:
            funding = [f.strip() for f in re.split(r"[;；]", funding_raw) if f.strip()]

        return Record(
            title=get("title"),
            authors=authors,
            year=year,
            journal=get("journal"),
            volume=get("volume") or None,
            issue=get("issue") or None,
            pages=get("pages") or None,
            doi=get("doi") or None,
            abstract=get("abstract") or None,
            keywords=keywords,
            doc_type=doc_type,
            language="zh",
            institutions=institutions,
            funding=funding,
            citation_count=citation_count,
            source_database="CNKI",
            source_file=source_file,
        )

    def _parse_authors(self, raw: str) -> list[Author]:
        """Parse CNKI author field.

        CNKI separates multiple authors with ";" or "；".
        """
        if not raw:
            return []

        names = re.split(r"[;；]", raw)
        authors = []
        for i, name in enumerate(names):
            name = name.strip()
            if not name:
                continue
            # CNKI format is typically surname+given in Chinese characters,
            # but we keep full_name for simplicity
            authors.append(Author(full_name=name, order=i + 1))
        return authors

    def _parse_institutions(self, raw: str) -> list[Institution]:
        """Parse institution field."""
        if not raw:
            return []

        inst_names = re.split(r"[;；]", raw)
        return [
            Institution(name=name.strip())
            for name in inst_names
            if name.strip()
        ]

    def _parse_keywords(self, raw: str) -> list[str]:
        """Parse keywords separated by various delimiters."""
        if not raw:
            return []

        # CNKI uses ";;", ";", "；", or spaces as separators
        # First try splitting by ";;"
        if ";;" in raw:
            parts = raw.split(";;")
        else:
            parts = re.split(r"[;；]", raw)

        return [k.strip() for k in parts if k.strip()]

    def _parse_doc_type(self, raw: str) -> DocType:
        """Map CNKI doc type strings to DocType enum."""
        if not raw:
            return DocType.UNKNOWN

        raw_lower = raw.strip().lower()

        type_map: dict[str, DocType] = {
            "期刊论文": DocType.ARTICLE,
            "期刊": DocType.ARTICLE,
            "article": DocType.ARTICLE,
            "综述": DocType.REVIEW,
            "review": DocType.REVIEW,
            "会议论文": DocType.CONFERENCE,
            "会议": DocType.CONFERENCE,
            "conference": DocType.CONFERENCE,
            "学位论文": DocType.THESIS,
            "博士": DocType.THESIS,
            "硕士": DocType.THESIS,
            "thesis": DocType.THESIS,
            "专利": DocType.PATENT,
            "patent": DocType.PATENT,
            "图书": DocType.BOOK,
            "book": DocType.BOOK,
        }

        for key, dtype in type_map.items():
            if key in raw_lower:
                return dtype

        return DocType.OTHER
