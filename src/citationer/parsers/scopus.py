"""Scopus CSV/Excel export parser."""

from __future__ import annotations

from pathlib import Path

from citationer.models.record import Author, DocType, Institution, Record
from citationer.parsers.base import BaseParser


class ScopusParser(BaseParser):
    """Parser for Scopus CSV and Excel export files.

    Scopus exports contain characteristic columns such as
    ``Authors``, ``Title``, ``Year``, ``Source title``, ``DOI``, etc.
    Detection checks for the presence of these column headers.
    """

    SCOPUS_MARKERS: list[str] = [
        "Authors",
        "Title",
        "Year",
        "Source title",
        "DOI",
    ]

    # Column header → Record field mapping
    COLUMN_MAP: dict[str, str] = {
        "Authors": "authors_raw",
        "Author(s) ID": "author_ids",
        "Title": "title",
        "Year": "year_raw",
        "Source title": "journal",
        "Volume": "volume",
        "Issue": "issue",
        "Pages": "pages",
        "Article Number": "art_no",
        "DOI": "doi",
        "Abstract": "abstract",
        "Author Keywords": "keywords_raw",
        "Index Keywords": "index_kw",
        "Language of Original Document": "language",
        "Document Type": "doc_type_raw",
        "Cited by": "citation_count",
        "References": "references_raw",
        "ISSN": "issn",
        "Funding Details": "funding_raw",
    }

    def __init__(self, encoding: str = "utf-8-sig") -> None:
        self._encoding = encoding

    @property
    def source_name(self) -> str:
        return "Scopus"

    def detect(self, filepath: Path) -> bool:
        suffix = filepath.suffix.lower()
        if suffix not in (".csv", ".xlsx"):
            return False

        try:
            headers = self._read_headers(filepath)
            if headers is None:
                return False
            header_text = " ".join(headers)
            match_count = sum(1 for m in self.SCOPUS_MARKERS if m in header_text)
            return match_count >= 3
        except Exception:
            return False

    def _read_headers(self, filepath: Path) -> list[str] | None:
        suffix = filepath.suffix.lower()
        if suffix == ".csv":
            return self._read_csv_headers(filepath)
        return self._read_xlsx_headers(filepath)

    def _read_csv_headers(self, filepath: Path) -> list[str] | None:
        import csv
        with open(filepath, encoding=self._encoding, errors="ignore") as f:
            reader = csv.reader(f)
            row = next(reader, None)
            if row:
                return [str(c).strip() for c in row]
        return None

    def _read_xlsx_headers(self, filepath: Path) -> list[str] | None:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        try:
            ws = wb.active
            if ws is None:
                return None
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if row:
                return [str(c).strip() if c else "" for c in row]
            return None
        finally:
            wb.close()

    def parse(self, filepath: Path) -> list[Record]:
        suffix = filepath.suffix.lower()
        if suffix == ".csv":
            return self._parse_csv(filepath)
        return self._parse_xlsx(filepath)

    # ------------------------------------------------------------------
    # CSV
    # ------------------------------------------------------------------

    def _parse_csv(self, filepath: Path) -> list[Record]:
        import csv
        import io

        with open(filepath, encoding=self._encoding, errors="ignore") as f:
            content = f.read()

        reader = csv.DictReader(io.StringIO(content))
        if reader.fieldnames is None:
            return []

        col_map = self._build_col_index(list(reader.fieldnames))
        records: list[Record] = []

        for row in reader:
            if all(not (v or "").strip() for v in row.values()):
                continue
            records.append(self._row_to_record(row, col_map, filepath.name))

        return records

    # ------------------------------------------------------------------
    # XLSX
    # ------------------------------------------------------------------

    def _parse_xlsx(self, filepath: Path) -> list[Record]:
        import openpyxl

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        try:
            ws = wb.active
            if ws is None:
                return []
            rows_iter = ws.iter_rows(values_only=True)
            first = next(rows_iter, None)
            if first is None:
                return []
            headers = [str(c).strip() if c else "" for c in first]
            col_map = self._build_col_index(headers)

            records: list[Record] = []
            for row in rows_iter:
                if all(c is None or str(c).strip() == "" for c in row):
                    continue
                row_dict = {
                    h: str(row[i]).strip() if i < len(row) and row[i] is not None else ""
                    for i, h in enumerate(headers)
                }
                records.append(self._row_to_record(row_dict, col_map, filepath.name))
            return records
        finally:
            wb.close()

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _build_col_index(self, headers: list[str]) -> dict[str, int]:
        """Map column names to field indices."""
        index: dict[str, int] = {}
        for i, h in enumerate(headers):
            h_clean = h.strip()
            for marker, field in self.COLUMN_MAP.items():
                if marker.lower() == h_clean.lower():
                    index[field] = i
                    break
        return index

    def _row_to_record(
        self,
        row: dict[str, str],
        col_map: dict[str, int],
        source_file: str,
    ) -> Record:
        def _val(field: str) -> str:
            # Scan row dict for matching field
            for k, v in row.items():
                for marker, f in self.COLUMN_MAP.items():
                    if f == field and marker.lower() == k.strip().lower():
                        return (v or "").strip()
            return ""

        # Parse year
        year: int | None = None
        year_str = _val("year_raw")
        if year_str:
            import re
            m = re.search(r"(\d{4})", year_str)
            if m:
                try:
                    year = int(m.group(1))
                except ValueError:
                    pass

        # Parse authors
        authors = self._parse_authors(_val("authors_raw"))

        # Parse keywords
        keywords: list[str] = []
        kw_str = _val("keywords_raw")
        if kw_str:
            keywords = [k.strip() for k in kw_str.split(";") if k.strip()]
        idx_str = _val("index_kw")
        if idx_str:
            keywords.extend(k.strip() for k in idx_str.split(";") if k.strip())

        # Parse institutions (from affiliation field or Author IDs)
        institutions: list[Institution] = []

        # Parse doc type
        doc_type = self._parse_doc_type(_val("doc_type_raw"))

        # Parse citation count
        citation_count: int | None = None
        cite_str = _val("citation_count")
        if cite_str:
            try:
                citation_count = int(cite_str)
            except ValueError:
                pass

        # Parse references
        refs_str = _val("references_raw")
        references: list[str] | None = None
        if refs_str:
            references = [r.strip() for r in refs_str.split(";") if r.strip()]

        # Parse funding
        fund_str = _val("funding_raw")
        funding: list[str] | None = None
        if fund_str:
            funding = [f.strip() for f in fund_str.split(";") if f.strip()]

        # Pages
        pages = _val("pages") or _val("art_no") or None

        return Record(
            title=_val("title"),
            authors=authors,
            year=year,
            journal=_val("journal") or None,
            volume=_val("volume") or None,
            issue=_val("issue") or None,
            pages=pages,
            doi=_val("doi") or None,
            issn=_val("issn") or None,
            abstract=_val("abstract") or None,
            keywords=keywords,
            doc_type=doc_type,
            language=_val("language") or "en",
            institutions=institutions,
            funding=funding,
            citation_count=citation_count,
            references=references,
            source_database="Scopus",
            source_file=source_file,
        )

    @staticmethod
    def _parse_authors(raw: str) -> list[Author]:
        if not raw:
            return []
        names = [n.strip() for n in raw.split(";") if n.strip()]
        authors = []
        for i, name in enumerate(names):
            surname, given = name, None
            if "," in name:
                parts = name.split(",", 1)
                surname = parts[0].strip()
                given = parts[1].strip() if len(parts) > 1 else None
            authors.append(Author(
                full_name=name,
                surname=surname,
                given_name=given,
                order=i + 1,
            ))
        return authors

    @staticmethod
    def _parse_doc_type(raw: str) -> DocType:
        raw_lower = raw.strip().lower()
        type_map: list[tuple[str, DocType]] = [
            ("conference paper", DocType.CONFERENCE),
            ("conference proceeding", DocType.CONFERENCE),
            ("book chapter", DocType.BOOK_CHAPTER),
            ("book", DocType.BOOK),
            ("article", DocType.ARTICLE),
            ("review", DocType.REVIEW),
            ("editorial", DocType.OTHER),
            ("letter", DocType.OTHER),
            ("note", DocType.OTHER),
            ("erratum", DocType.OTHER),
        ]
        for key, dt in type_map:
            if key in raw_lower:
                return dt
        return DocType.UNKNOWN
