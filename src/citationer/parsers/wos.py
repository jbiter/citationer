"""Web of Science (WoS) text/Excel export parser."""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path

# openpyxl is imported lazily in WosExcelParser methods for .xlsx handling
from citationer.models.record import Author, DocType, Institution, Record
from citationer.parsers.base import BaseParser


class WosTextParser(BaseParser):
    """Parser for Web of Science plain-text tagged export files (.txt, .ciw).

    WoS exports in "Full Record and Cited References" format use
    two-character field tags (e.g. "PT", "AU", "TI", "SO", ...).
    Continuation lines are indented with 3 spaces.
    Records are terminated by "ER".
    The file begins with "FN Clarivate Analytics Web of Science" and "VR 1.0".
    """

    # WoS two-character field tags
    WOS_TAGS: set[str] = {
        "PT", "AU", "AF", "TI", "SO", "LA", "DT", "DE", "ID",
        "AB", "C1", "RP", "EM", "FU", "FX", "CR", "NR", "TC",
        "Z9", "PU", "PI", "PA", "SN", "EI", "BN", "J9", "JI",
        "PD", "PY", "VL", "IS", "BP", "EP", "AR", "DI", "UT",
        "ER", "EF", "PM", "WC", "SC", "GA", "U1", "U2",
        "DA", "PG", "SU", "SI", "MA", "D2",
    }

    def __init__(self, encoding: str = "utf-8-sig") -> None:
        self._encoding = encoding

    @property
    def source_name(self) -> str:
        return "WoS"

    def detect(self, filepath: Path) -> bool:
        """Detect WoS tagged text export by checking for characteristic markers."""
        suffix = filepath.suffix.lower()
        if suffix not in (".txt", ".ciw"):
            return False

        try:
            with open(filepath, encoding="utf-8-sig", errors="ignore") as f:
                first_line = f.readline().strip()

            # WoS tagged format starts with "FN Clarivate..." or "PT J" (tag + space + content)
            # Exclude tab-delimited files (which have "\t" after the tag)
            if "\t" in first_line:
                return False

            return (
                first_line.startswith("FN ")
                or (first_line[:2] in self.WOS_TAGS and first_line[2:3] == " ")
            )
        except Exception:
            return False

    def parse(self, filepath: Path) -> list[Record]:
        """Parse a WoS tagged text export into Record objects."""
        records: list[Record] = []
        current: dict[str, str] = {}
        current_tag: str | None = None

        with open(filepath, encoding="utf-8-sig", errors="ignore") as f:
            for line in f:
                line = line.rstrip("\n\r")

                # Skip file header/footer lines
                if line.startswith("FN ") or line.startswith("VR "):
                    continue
                if line == "EF":
                    break

                # End of record — "ER" alone (may or may not have trailing content)
                if line == "ER" or line.startswith("ER ") and len(line) <= 3:
                    if current:
                        records.append(self._dict_to_record(current, filepath.name))
                        current = {}
                        current_tag = None
                    continue

                # Check if line starts with a known two-character WoS tag
                # Format: "XX value" where XX is a 2-char tag followed by a space
                if (
                    len(line) >= 3
                    and line[:2] in self.WOS_TAGS
                    and line[2:3] == " "
                ):
                    tag = line[:2]
                    # Skip record boundary tags
                    if tag in ("ER", "EF"):
                        continue
                    current_tag = tag
                    current[current_tag] = line[3:].strip()
                elif current_tag and line.startswith("   "):
                    # Continuation line: 3+ spaces of indentation
                    continuation = line.strip()
                    if continuation:
                        # For author fields, use "; " separator to distinguish
                        # multi-author lines from wrapped text
                        sep = "; " if current_tag in ("AU", "AF", "C1") else " "
                        current[current_tag] += sep + continuation
                elif current_tag and line.strip():
                    # Some WoS exports wrap without indentation
                    continuation = line.strip()
                    current[current_tag] += " " + continuation

        # Handle last record if no ER terminator
        if current:
            records.append(self._dict_to_record(current, filepath.name))

        return records

    def _dict_to_record(self, fields: dict[str, str], source_file: str) -> Record:
        """Convert a WoS field dictionary to a Record."""

        def get(tag: str) -> str:
            return fields.get(tag, "")

        # Parse year
        year: int | None = None
        year_str = get("PY")
        if year_str:
            try:
                year = int(year_str.strip())
            except ValueError:
                pass

        # Parse authors (AU = abbreviated, AF = full names)
        authors = self._parse_authors(get("AU"), get("AF"))

        # Parse institutions from C1 field
        institutions = self._parse_institutions(get("C1"))

        # Parse keywords (DE = author keywords, ID = Keywords Plus)
        keywords: list[str] = []
        de_str = get("DE")
        id_str = get("ID")
        if de_str:
            keywords.extend(self._parse_semicolon_list(de_str))
        if id_str:
            keywords.extend(self._parse_semicolon_list(id_str))

        # Parse doc type
        doc_type = self._parse_doc_type(get("DT"))

        # Parse citation count (TC = Web of Science Core Collection, Z9 = Total)
        citation_count: int | None = None
        tc_str = get("TC") or get("Z9")
        if tc_str.strip():
            try:
                citation_count = int(tc_str.strip())
            except ValueError:
                pass

        # References count
        references: list[str] | None = None
        cr_str = get("CR")
        if cr_str:
            refs = re.split(r";\s*(?=[A-Z])", cr_str)
            references = [r.strip() for r in refs if r.strip()]

        doi = get("DI") or None

        # Pages: use BP-EP, or AR (article number), or PG
        pages = self._format_pages(get("BP"), get("EP"))
        if not pages:
            pages = get("AR") or None
        if not pages:
            pages = get("PG") or None

        return Record(
            title=get("TI"),
            title_en=get("TI"),  # WoS titles are in English
            authors=authors,
            year=year,
            journal=get("SO") or get("JI") or get("J9"),
            volume=get("VL") or None,
            issue=get("IS") or None,
            pages=pages,
            doi=doi,
            issn=get("SN") or None,
            abstract=get("AB") or None,
            keywords=keywords,
            doc_type=doc_type,
            language=get("LA") or "en",
            institutions=institutions,
            funding=self._parse_funding(get("FU"), get("FX")),
            citation_count=citation_count,
            references=references,
            source_database="WoS",
            source_file=source_file,
        )

    def _parse_authors(self, au_field: str, af_field: str) -> list[Author]:
        """Parse WoS author fields.

        AU = abbreviated names (e.g. "Smith, J; Jones, M")
        AF = full names with given names spelled out
        """
        raw = af_field or au_field
        if not raw:
            return []

        names = self._parse_semicolon_list(raw)
        authors: list[Author] = []
        for i, name in enumerate(names):
            name = name.strip()
            if not name:
                continue

            # Parse "Surname, GivenName" format
            if "," in name:
                parts = name.split(",", 1)
                surname = parts[0].strip()
                given = parts[1].strip() if len(parts) > 1 else None
            else:
                surname = name
                given = None

            authors.append(
                Author(
                    full_name=name,
                    surname=surname,
                    given_name=given,
                    order=i + 1,
                )
            )
        return authors

    def _parse_institutions(self, c1_field: str) -> list[Institution]:
        """Parse WoS C1 (author address) field.

        Handles both WoS export formats:

        *Tagged text*:  ``[Smith, John] Harvard Univ, Cambridge, MA 02138 USA``
        *Excel*:        ``Smith, John; Harvard Univ, Cambridge, MA 02138 USA``

        Multiple addresses are separated by newlines (tagged) or
        ``. `` (dot-space, Excel).  The parsed result deduplicates
        institution names (case-insensitive).

        """
        if not c1_field:
            return []

        # ── Normalise & split into per-author address entries ──────────
        text = c1_field.strip()

        # Strategy: the most reliable separator between different
        # author-address groups is a newline OR a ``[`` marker (lookahead).
        # Excel exports typically use newlines; tagged exports use ``[``.
        entries = re.split(r"\n|(?=\[)", text)
        entries = [e.strip().rstrip(".") for e in entries if e.strip()]

        # If splitting didn't produce multiple entries, try semicolons
        if len(entries) == 1:
            # Excel may put everything on one line separated by "; "
            # but only split when we see author-like patterns: "Name, Init.;"
            entries = re.split(r";\s*(?=[A-Z][a-z]+,\s+[A-Z])", entries[0])
            entries = [e.strip() for e in entries if e.strip()]

        institutions: list[Institution] = []
        seen: set[str] = set()

        for entry in entries:
            entry = entry.strip().rstrip(".")
            if not entry:
                continue

            # ── Remove author information ──────────────────────────
            # Format A:  "[Surname, GivenName] Institution, City, Country"
            cleaned = re.sub(r"^\s*\[[^\]]*\]\s*", "", entry).strip()

            # Format B:  "Surname, GivenName; Institution, City, Country"
            # Author part: one or two words separated by comma, ending with ";"
            cleaned = re.sub(
                r"^[A-Z][a-z]+(?:-[A-Z][a-z]+)?,\s+[A-Z][a-z.]*(?:\s+[A-Z][a-z.]*)?;\s*",
                "", cleaned,
            ).strip()

            if not cleaned or cleaned.lower().startswith("reprint"):
                continue

            # ── Extract institution name ───────────────────────────
            tokens = [t.strip() for t in cleaned.split(",")]

            # Common institution-indicating keywords
            _inst_kw = {
                "univ", "inst", "coll", "sch", "ctr", "lab", "dept",
                "hosp", "acad", "corp", "inc", "ltd", "gmbh", "sa",
                "res", "found", "council", "minist", "agcy", "authority",
                "natl", "nation", "european", "chinese", "japan",
                "us", "usa", "peking", "tsing", "zhejiang", "fudan",
                "shanghai", "beijing", "nanjing", "wuhan", "harbin",
            }

            inst_name = ""
            # Walk the first few comma-separated tokens; pick the first
            # one that looks like a real institution name.
            for token in tokens[:4]:
                token_stripped = token.strip()
                token_lower = token_stripped.lower()
                # Single short words (likely leftover author surnames) → skip
                words = token_stripped.split()
                if len(words) == 1 and len(token_stripped) <= 10 and not any(
                    kw in token_lower for kw in _inst_kw
                ):
                    continue
                # Numeric / zip-code tokens → skip
                if token_stripped.replace("-", "").isdigit():
                    continue
                inst_name = token_stripped
                break

            if not inst_name:
                continue

            # ── Extract country ────────────────────────────────────
            country: str | None = None
            if len(tokens) > 1:
                last = tokens[-1].strip().rstrip(".")
                if (
                    re.match(r"^[A-Za-z\s]{2,30}$", last)
                    and not any(c.isdigit() for c in last)
                ):
                    country = last

            dedup_key = inst_name.lower()
            if dedup_key and dedup_key not in seen:
                seen.add(dedup_key)
                institutions.append(Institution(name=inst_name, country=country))

        return institutions

    def _parse_doc_type(self, raw: str) -> DocType:
        """Map WoS document type to DocType enum."""
        if not raw:
            return DocType.UNKNOWN

        raw_lower = raw.strip().lower()

        type_map: dict[str, DocType] = {
            "article": DocType.ARTICLE,
            "review": DocType.REVIEW,
            "proceedings paper": DocType.CONFERENCE,
            "meeting abstract": DocType.CONFERENCE,
            "editorial material": DocType.ARTICLE,
            "letter": DocType.ARTICLE,
            "book review": DocType.REVIEW,
            "book": DocType.BOOK,
            "book chapter": DocType.BOOK_CHAPTER,
            "patent": DocType.PATENT,
            "thesis": DocType.THESIS,
            "note": DocType.OTHER,
            "correction": DocType.OTHER,
            "retracted publication": DocType.OTHER,
            "news item": DocType.OTHER,
            "biographical-item": DocType.OTHER,
        }

        return type_map.get(raw_lower, DocType.OTHER)

    @staticmethod
    def _parse_semicolon_list(raw: str) -> list[str]:
        """Split a WoS semicolon-delimited field."""
        return [s.strip() for s in raw.split(";") if s.strip()]

    @staticmethod
    def _format_pages(bp: str, ep: str) -> str | None:
        """Combine begin page and end page."""
        bp = bp.strip()
        ep = ep.strip()
        if bp and ep:
            return f"{bp}-{ep}"
        if bp:
            return bp
        return None

    @staticmethod
    def _parse_funding(fu_field: str, fx_field: str) -> list[str] | None:
        """Parse funding information."""
        funding: list[str] = []
        if fu_field:
            funding.extend(
                s.strip() for s in fu_field.split(";") if s.strip()
            )
        if fx_field:
            funding.extend(
                s.strip() for s in fx_field.split(";") if s.strip()
            )
        return funding if funding else None


class WosTabDelimitedParser(BaseParser):
    """Parser for Web of Science tab-delimited export files.

    WoS also exports in a tab-separated format with column headers.
    The first row contains field tag headers (PT, AU, TI, etc.).
    """

    # Field tag to column header mapping (lowercased for matching)
    TAG_MAPPING: dict[str, str] = {
        "pt": "PT",
        "au": "AU",
        "af": "AF",
        "ti": "TI",
        "so": "SO",
        "la": "LA",
        "dt": "DT",
        "de": "DE",
        "id": "ID",
        "ab": "AB",
        "c1": "C1",
        "rp": "RP",
        "em": "EM",
        "fu": "FU",
        "fx": "FX",
        "cr": "CR",
        "nr": "NR",
        "tc": "TC",
        "z9": "Z9",
        "pu": "PU",
        "pi": "PI",
        "pa": "PA",
        "sn": "SN",
        "ei": "EI",
        "j9": "J9",
        "ji": "JI",
        "pd": "PD",
        "py": "PY",
        "vl": "VL",
        "is": "IS",
        "bp": "BP",
        "ep": "EP",
        "ar": "AR",
        "di": "DI",
        "ut": "UT",
        "ga": "GA",
        "wc": "WC",
        "sc": "SC",
    }

    def __init__(self, encoding: str = "utf-8-sig") -> None:
        self._encoding = encoding

    @property
    def source_name(self) -> str:
        return "WoS"

    def detect(self, filepath: Path) -> bool:
        """Detect WoS tab-delimited export by checking for tab-separated header."""
        suffix = filepath.suffix.lower()
        if suffix not in (".txt", ".tsv", ".csv"):
            return False

        # First, make sure it's NOT the tagged format
        try:
            with open(filepath, encoding="utf-8-sig", errors="ignore") as f:
                first_line = f.readline().strip()
            # If it starts with FN or PT (tagged format), skip
            if first_line.startswith("FN ") or first_line.startswith("PT "):
                return False
            # Tab-delimited has "PT\tAU\t..." as first line
            if "\t" in first_line and any(
                tag.lower() in first_line.lower() for tag in ["PT", "AU", "TI"]
            ):
                return True
        except Exception:
            pass

        return False

    def parse(self, filepath: Path) -> list[Record]:
        """Parse a WoS tab-delimited export into Record objects."""
        with open(filepath, encoding="utf-8-sig", errors="ignore") as f:
            content = f.read()

        # Use csv.DictReader with tab delimiter
        reader = csv.DictReader(
            io.StringIO(content),
            delimiter="\t",
            quoting=csv.QUOTE_NONE,
        )

        # Map column headers to WoS tags
        header_map: dict[str, str] = {}
        if reader.fieldnames:
            for col in reader.fieldnames:
                col_clean = col.strip().lower()
                if col_clean in self.TAG_MAPPING:
                    header_map[col] = self.TAG_MAPPING[col_clean]

        # Use the text parser's _dict_to_record for conversion
        text_parser = WosTextParser()
        records: list[Record] = []

        for row in reader:
            # Convert column names to two-char tags
            fields: dict[str, str] = {}
            for col, value in row.items():
                if col in header_map:
                    tag = header_map[col]
                    fields[tag] = value.strip()

            if fields:
                records.append(
                    text_parser._dict_to_record(fields, filepath.name)
                )

        return records


class WosExcelParser(BaseParser):
    """Parser for Web of Science Excel export files (.xlsx and .xls)."""

    # WoS export header markers for detection
    WOS_MARKERS: list[str] = [
        "Publication Type",
        "Article Title",
        "Authors",
        "Source Title",
        "Times Cited",
    ]

    # Column header → WoS field tag mapping
    FIELD_MARKERS: dict[str, str] = {
        "Publication Type": "DT",
        "Document Type": "DT",
        "Article Title": "TI",
        "Authors": "AU",
        "Author Full Names": "AF",
        "Source Title": "SO",
        "Journal": "SO",
        "Abstract": "AB",
        "Author Keywords": "DE",
        "Keywords Plus": "ID",
        "Addresses": "C1",
        "Correspondence Address": "RP",
        "ISSN": "SN",
        "DOI": "DI",
        "Publication Year": "PY",
        "Volume": "VL",
        "Issue": "IS",
        "Start Page": "BP",
        "End Page": "EP",
        "Article Number": "AR",
        "Times Cited, WoS Core": "TC",
        "Times Cited, All Databases": "Z9",
        "Cited References": "CR",
        "Language": "LA",
        "Funding Orgs": "FU",
        "Funding Text": "FX",
    }

    def __init__(self, encoding: str = "utf-8") -> None:
        self._encoding = encoding

    @property
    def source_name(self) -> str:
        return "WoS"

    def detect(self, filepath: Path) -> bool:
        """Detect WoS Excel export by checking header row."""
        if filepath.suffix.lower() not in (".xlsx", ".xls"):
            return False

        try:
            # First confirm it's NOT a CNKI export
            from citationer.parsers.cnki import CnkiExcelParser

            cnki = CnkiExcelParser()
            if cnki.detect(filepath):
                return False
        except Exception:
            pass

        # Try reading headers with the appropriate library
        headers = self._read_headers(filepath)
        if headers is None:
            return False

        header_text = " ".join(headers)
        match_count = sum(1 for m in self.WOS_MARKERS if m in header_text)
        return match_count >= 2

    def parse(self, filepath: Path) -> list[Record]:
        """Parse a WoS Excel export (.xlsx or .xls)."""
        suffix = filepath.suffix.lower()

        if suffix == ".xls":
            return self._parse_xls(filepath)
        else:
            return self._parse_xlsx(filepath)

    # ------------------------------------------------------------------
    # Header reading
    # ------------------------------------------------------------------

    def _read_headers(self, filepath: Path) -> list[str] | None:
        """Read the header row, dispatching to the right library."""
        suffix = filepath.suffix.lower()
        try:
            if suffix == ".xls":
                return self._read_headers_xls(filepath)
            else:
                return self._read_headers_xlsx(filepath)
        except Exception:
            return None

    def _read_headers_xlsx(self, filepath: Path) -> list[str] | None:
        """Read header row from .xlsx using openpyxl."""
        import openpyxl

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        try:
            ws = wb.active
            if ws is None:
                return None
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if row:
                return [str(c).strip() if c else "" for c in row]
            return None
        finally:
            wb.close()

    def _read_headers_xls(self, filepath: Path) -> list[str] | None:
        """Read header row from .xls using xlrd."""
        import xlrd

        wb = xlrd.open_workbook(str(filepath))
        try:
            ws = wb.sheet_by_index(0)
            if ws.nrows == 0:
                return None
            return [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
        finally:
            pass  # xlrd auto-closes on garbage collection

    # ------------------------------------------------------------------
    # Parsing
    # ------------------------------------------------------------------

    def _parse_xlsx(self, filepath: Path) -> list[Record]:
        """Parse .xlsx file with openpyxl."""
        import openpyxl

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        try:
            ws = wb.active
            if ws is None:
                return []

            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

        if len(rows) < 2:
            return []

        headers = [str(c).strip() if c else "" for c in rows[0]]
        return self._rows_to_records(headers, rows[1:], filepath.name)

    def _parse_xls(self, filepath: Path) -> list[Record]:
        """Parse .xls file with xlrd."""
        import xlrd

        wb = xlrd.open_workbook(str(filepath))
        try:
            ws = wb.sheet_by_index(0)
            if ws.nrows < 2:
                return []

            headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
            rows = [
                tuple(ws.cell_value(r, c) for c in range(ws.ncols))
                for r in range(1, ws.nrows)
            ]
        finally:
            pass  # xlrd auto-closes on garbage collection

        return self._rows_to_records(headers, rows, filepath.name)

    def _build_col_map(self, headers: list[str]) -> dict[str, int]:
        """Build column index from header names to WoS field tags.

        Uses exact match first (case-insensitive), then falls back to
        substring match. Exact matches are never overwritten by later
        substring matches, preventing false positives like "Authors"
        matching "Group Authors" instead of the real "Authors" column.
        """
        col_map: dict[str, int] = {}
        exact_matched: set[str] = set()

        for i, h in enumerate(headers):
            h_lower = h.lower()

            # Phase 1: exact match (case-insensitive)
            for marker, tag in self.FIELD_MARKERS.items():
                if marker.lower() == h_lower:
                    col_map[tag] = i
                    exact_matched.add(tag)
                    break
            else:
                # Phase 2: substring match — only if tag not already
                # found via exact match in an earlier column
                for marker, tag in self.FIELD_MARKERS.items():
                    if tag not in exact_matched and marker.lower() in h_lower:
                        col_map[tag] = i
                        break

        return col_map

    def _rows_to_records(
        self,
        headers: list[str],
        rows: list[tuple],
        source_file: str,
    ) -> list[Record]:
        """Convert rows + headers to Record objects."""
        col_map = self._build_col_map(headers)

        text_parser = WosTextParser()
        records: list[Record] = []

        for row in rows:
            # Skip empty rows
            if all(
                c is None or (isinstance(c, float) and c == 0.0) or str(c).strip() == ""
                for c in row
            ):
                continue

            fields: dict[str, str] = {}
            for tag, idx in col_map.items():
                if idx < len(row):
                    val = row[idx]
                    if val is not None:
                        # xlrd returns floats for numeric cells — convert cleanly
                        if isinstance(val, float):
                            if val == int(val):
                                val = str(int(val))
                            else:
                                val = str(val)
                        else:
                            val = str(val).strip()
                        if val:
                            fields[tag] = val

            if fields:
                records.append(
                    text_parser._dict_to_record(fields, source_file)
                )

        return records
